from openpyxl import load_workbook
from dataclasses import dataclass
from typing import List

@dataclass
class Article:
    name: str
    quantity: int

@dataclass # classe des articles non trouvés
class Error:
    article: Article
    client: str
    
LOWEST_CLIENT_ROW = 4
HIGHEST_CLIENT_ROW = 66

CENTRAL_WORKBOOK = "Commandes.xlsx"
OUTPUT_WORKBOOK = "Commandes2.xlsx"

def add_article(sheet, client, article: Article, errors: List[Error]):
    for row in range(LOWEST_CLIENT_ROW, HIGHEST_CLIENT_ROW):

        order_name = sheet.cell(row=row, column=1).value
        order_quantity = sheet.cell(row=row, column=11).value

        if order_name == article.name:
            if order_quantity is None:
                order_quantity = 0
            
            sheet.cell(row=row, column=11).value = order_quantity + article.quantity
            
            return  # on sort dès qu'on a trouvé
    print_missing_article(article)
    add_missing_article(client, article, errors)
    

def print_missing_article(missing_article: Article):
    print("!!! ATTENTION : l'article " + missing_article.name + " n'a pas été trouvé dans " + CENTRAL_WORKBOOK + " !!!")
    print("Une cellule a peut être été modifiée.")
    input("Appuyez sur Entrée...")

def add_missing_article(client, article: Article, errors: List[Error]):
    errors.append(Error(article=article, client=client))

def print_errors(errors: List[Error]):
    if errors:
        print("!!! CERTAINS ARTICLES N'ONT PAS PU ETRE AJOUTES !!!")
        for error in errors:
            msg = (
                f"{error.article.quantity} {error.article.name} pour le client {error.client} "
                f"n'ont pas pu être ajoutés à {CENTRAL_WORKBOOK} "
                f"car ils n'ont pas été trouvés."
            )
            print(msg)
    if not errors:
        print("Aucune erreur.")

def read_articles(sheet):
    
    articles = []

    # Cherche les articles dont la quantité n'est pas nulle
    for name, quantity in sheet.iter_rows(
        min_row=LOWEST_CLIENT_ROW,
        max_row=HIGHEST_CLIENT_ROW+1,
        min_col=1,
        max_col=2,
        values_only=True
    ):
        if quantity is not None and quantity != "":
            try:
                quantity_int = int(quantity)
            except (TypeError, ValueError):
                continue  # ou gérer autrement

            current_article = Article(name, quantity_int)
            articles.append(current_article)
            

    return articles
            
        
def main():
    
    # Charger les fichiers
    client_workbook = load_workbook("Gamm vert.xlsx")
    client_sheet = client_workbook.active

    orders_workbook = load_workbook(CENTRAL_WORKBOOK)
    orders_sheet = orders_workbook.active

    errors = []
    articles = read_articles(client_sheet)


    for article in articles:
        print(article)
        add_article(orders_sheet, "Gamm vert", article, errors)
    
    orders_workbook.save(OUTPUT_WORKBOOK)


    print_errors(errors)
    print("Done")

if __name__ == "__main__":
    main()