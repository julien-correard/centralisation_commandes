from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from typing import List

from models import Article, Error

from config import CLIENT_NAME_ROW, CLIENT_NAME_COLUMN, LOWEST_CLIENT_ROW, HIGHEST_CLIENT_ROW
    


def get_client(sheet, client_filename):
    value = sheet.cell(row=CLIENT_NAME_ROW, column=CLIENT_NAME_COLUMN).value

    if value is None or value == "":
        # Traduction de 1 3 par exemple à C1
        col_letter = get_column_letter(CLIENT_NAME_COLUMN)
        cell_ref = f"{col_letter}{CLIENT_NAME_ROW}"

        raise ValueError(
            f"Nom du client introuvable dans la cellule "
            f"{cell_ref} "
            f"du fichier {client_filename}.\n"
            f"Veuillez corriger le fichier et relancer le script."
        )

    return value

def read_articles(sheet):
    
    articles = []

    # Cherche les articles dont la quantité n'est pas nulle
    for name, quantity in sheet.iter_rows(
        min_row=LOWEST_CLIENT_ROW,
        max_row=HIGHEST_CLIENT_ROW+1,
        min_col=1,
        max_col=2,
        values_only=True
    ):
        if quantity is not None and quantity != "":
            try:
                quantity_int = int(quantity)
            except (TypeError, ValueError):
                continue  # ou gérer autrement

            current_article = Article(name, quantity_int)
            articles.append(current_article)
            

    return articles