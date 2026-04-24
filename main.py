from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from typing import List

from models import Article
from excel_writer import add_article
from excel_reader import get_client, read_articles
from client_mapping import load_client_map

from config import CLIENT_NAME_ROW, CLIENT_NAME_COLUMN, LOWEST_CLIENT_ROW, HIGHEST_CLIENT_ROW
from config import CENTRAL_WORKBOOK, OUTPUT_WORKBOOK

        
def main():
    try:
        # Charger les fichiers
        client_filename = "Gamm vert.xlsx"
        client_workbook = load_workbook(client_filename)
        client_sheet = client_workbook.active

        orders_workbook = load_workbook(CENTRAL_WORKBOOK)
        orders_sheet = orders_workbook.active

        client_map = load_client_map()

        articles = read_articles(client_sheet)
        client = get_client(client_sheet, client_filename)

        for article in articles:
            print(article)
            add_article(orders_sheet, client, article, client_map)
        
        orders_workbook.save(OUTPUT_WORKBOOK)
        print(f"Fichier {OUTPUT_WORKBOOK} enregistré.")

        print("Done")

    except ValueError as e:
        print("\n!!! ERREUR !!!\n")
        print(e)
        print("\nAucun fichier n'a été sauvegardé.")
        input("\nAppuyez sur Entrée pour quitter...")

if __name__ == "__main__":
    main()