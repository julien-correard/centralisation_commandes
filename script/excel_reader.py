from openpyxl.utils import get_column_letter

from cell_locker import check_if_cell_unlocked, check_if_sheet_locked
from excel_writer import save_workbook

from models import Article, Cell

def get_client(sheet, client_filename,config):

    row=config.client_name_row
    col=config.client_name_column
    
    value = sheet.cell(row=row, column=col).value

    if value is None or value == "":
        # Traduction de 1 3 par exemple à C1
        col_letter = get_column_letter(config.client_name_column)
        cell_ref = f"{col_letter}{config.client_name_row}"

        raise ValueError(
            f"Nom du client introuvable dans la cellule "
            f"{cell_ref} "
            f"du fichier {client_filename}.\n"
            f"Veuillez corriger le fichier et relancer le script."
        )
    return value

def read_articles(wb, sheet, client, filename, config):
    articles = []

    for row in range(config.lowest_client_row, config.highest_client_row + 1):

        col = 1
        while col <= sheet.max_column:

            name = sheet.cell(row=row, column=col).value
            quantity = sheet.cell(row=row, column=col + 1).value

            if name not in (None, ""):

                check_if_cell_unlocked(sheet, row, col + 1, config.protect_client_files)

                if quantity not in (None, ""):
                    try:
                        quantity_int = int(quantity)
                    except (TypeError, ValueError):
                        col_letter = get_column_letter(col+1)
                        raise ValueError(
                            f"Quantité non valide pour le client {client} :\n"
                            f"'{name}' à la ligne {row}, colonne {col_letter}: {quantity}\n"
                            f"Veuillez vérifier le fichier {filename} et réessayer."
                            )
                    articles.append(Article(name, quantity_int))

            col += 2
    check_if_sheet_locked(sheet, filename, config.protect_client_files) # Verrouillage des cellules après lecture ET SAUVEGARDE

    return articles