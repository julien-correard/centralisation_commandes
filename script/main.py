from openpyxl import load_workbook

from pathlib import Path

import sys

from excel_writer import add_article, save_workbook, clear_central_table
from excel_reader import get_client, read_articles
from client_mapping import load_client_map

from config_loader import load_config
from central_file_save import save_central_file, delete_temp_central_file, save_temp_backup_central_file

        
def main():
    try:
        config = None
        save_file = None
        if len(sys.argv) > 1:
            BASE_DIR = Path(sys.argv[1])
        else:
            BASE_DIR = Path(__file__).resolve().parents[1]
        
        try:
            config = load_config(BASE_DIR)
        except Exception as e:
            print(e)
            exit(1)
        
        client_map, client_list = load_client_map(config.fichier_correspondance_clients)

        try:
            orders_workbook = load_workbook(config.central_workbook)
        except Exception as e:
            raise ValueError(
                f"Impossible d'ouvrir le fichier Excel {config.central_workbook} : {e}"
            )
        orders_sheet = orders_workbook.active

        save_file = save_temp_backup_central_file(config)

        clear_central_table(orders_sheet, config)

        clients_folder = config.client_path

        files = list(clients_folder.glob("*.xlsx"))

        if not files:
            raise FileNotFoundError(
                f"Aucun fichier .xlsx trouvé dans le dossier {clients_folder}"
            )

        for file_path in files:

            print(f"Traitement du fichier {file_path} : ", end="")

            client_workbook = load_workbook(file_path)
            client_sheet = client_workbook.active

            client = get_client(client_sheet, file_path, config)
            articles = read_articles(client_workbook, client_sheet, client, file_path, config)

            for article in articles:
                add_article(orders_sheet, client, article, client_map, config)
            print("OK")

            client_list.discard(client) #Vérifier si tous les clients ont été traités

        save_workbook(orders_workbook, config.output_workbook)

        for missing_client in client_list:
            print(f"\n!!! Attention : le client {missing_client} n'a pas pu être traité, fichier .xlsx introuvable !!!\n"
                  f"Si ce client n'est plus en compte, veuillez le supprimer du fichier :"
                  f" {Path(config.fichier_correspondance_clients).relative_to(BASE_DIR)}")

        print(f"\nFichier {Path(config.output_workbook).relative_to(config.working_directory)} enregistré avec succès.\n")

        save_central_file(save_file)

        input("Appuyez sur entrée pour quitter...")

    except (ValueError, FileNotFoundError) as e:
        
        if save_file:
            delete_temp_central_file(save_file)

        output_file = config.output_workbook if config else "inconnu"
        message = (
        f"\n!!! ERREUR !!!\n\n"
        f"{e}\n\n"
        f"Le fichier {output_file} n'a pas été modifié.\n"
        f"\nAppuyez sur Entrée pour quitter..."
        )
        print(message)
        input()

if __name__ == "__main__":
    main()
